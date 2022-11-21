import openpyxl
import re
import datetime

def open_workbook(file):
    print('Открывается файл...')
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    max_rows = ws.max_row


    print('Поиск совпадений ...')
    regex = re.compile(r'\w*\+\w*')
    list=[]
    for rows in range(1, (max_rows + 1)):
        result = ws.cell(row=rows, column=3).value
        list.append(result)
    print(list)
    my_list = ''.join(list)
    findall = regex.findall(my_list)
    print(len(findall))


    print('Готово')
if __name__ == '__main__':
    open_workbook('Q:\Python\Scan_Excel\example2.xlsx')